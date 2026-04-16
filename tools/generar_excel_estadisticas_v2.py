#!/usr/bin/env python3
"""
Excel v2: Ejemplo simple e intuitivo de estadísticas básicas.
Caso: Tiempo de llegada de colectivos a la terminal.
¿Llegan a horario? ¿Quién es puntual y quién no?

Compatible con Excel 2016+.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Estilos ───
NAVY = "0F172A"
BLUE = "3B82F6"
TEAL = "0D9488"
ORANGE = "EA580C"
GREEN = "059669"
WHITE = "FFFFFF"
LIGHT = "F1F5F9"
MID = "E2E8F0"
RED = "DC2626"

hfont = Font(name="Calibri", size=11, bold=True, color=WHITE)
hfill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
tfont = Font(name="Calibri", size=14, bold=True, color=NAVY)
sfont = Font(name="Calibri", size=12, bold=True, color=BLUE)
nfont = Font(name="Calibri", size=11)
bfont = Font(name="Calibri", size=11, bold=True)
note = Font(name="Calibri", size=10, italic=True, color="64748B")
gfont = Font(name="Calibri", size=11, bold=True, color=GREEN)
ofont = Font(name="Calibri", size=11, bold=True, color=ORANGE)
rfont = Font(name="Calibri", size=11, bold=True, color=RED)
blfont = Font(name="Calibri", size=11, bold=True, color=BLUE)
border = Border(
    left=Side(style="thin", color=MID), right=Side(style="thin", color=MID),
    top=Side(style="thin", color=MID), bottom=Side(style="thin", color=MID),
)
ctr = Alignment(horizontal="center", vertical="center")
wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
alt = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
green_bg = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
orange_bg = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")
red_bg = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")


def header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = ctr
        cell.border = border


def style_rows(ws, r1, r2, cols):
    for r in range(r1, r2 + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = nfont
            cell.border = border
            cell.alignment = ctr
            if (r - r1) % 2 == 1:
                cell.fill = alt


def widths(ws, w_list):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════
# HOJA 1: "Llegadas a Terminal"
# Caso super simple: minutos de demora de colectivos
# ═══════════════════════════════════════════════════
ws = wb.active
ws.title = "Llegadas a Terminal"
ws.sheet_properties.tabColor = BLUE

ws.merge_cells("A1:D1")
ws["A1"].value = "¿Los colectivos llegan a horario a la terminal?"
ws["A1"].font = tfont

ws.merge_cells("A2:D2")
ws["A2"].value = "Minutos de demora en las últimas 15 llegadas de la Ruta 3. Negativo = llegó antes, 0 = puntual, positivo = llegó tarde."
ws["A2"].font = note
ws.row_dimensions[2].height = 30

for i, h in enumerate(["Llegada #", "Chofer", "Demora (min)", "Estado"], 1):
    ws.cell(row=4, column=i, value=h)
header(ws, 4, 4)

# Datos simples: minutos de demora
# Diseñados para que promedio ≈ 3 min, mediana = 2, moda = 2, con 2 outliers
llegadas = [
    (1, "García", 2, "A horario"),
    (2, "López", -1, "Llegó antes"),
    (3, "García", 3, "Leve demora"),
    (4, "Martínez", 2, "A horario"),
    (5, "López", 1, "A horario"),
    (6, "Fernández", 2, "A horario"),
    (7, "García", 4, "Leve demora"),
    (8, "Rodríguez", 2, "A horario"),
    (9, "López", 0, "Puntual"),
    (10, "Martínez", 3, "Leve demora"),
    (11, "Fernández", 2, "A horario"),
    (12, "García", 1, "A horario"),
    (13, "Rodríguez", 25, "Muy tarde"),     # OUTLIER
    (14, "Martínez", 2, "A horario"),
    (15, "Fernández", 35, "Muy tarde"),     # OUTLIER
]

for i, (num, chofer, demora, estado) in enumerate(llegadas, 5):
    ws.cell(row=i, column=1, value=num)
    ws.cell(row=i, column=2, value=chofer)
    ws.cell(row=i, column=3, value=demora)
    ws.cell(row=i, column=4, value=estado)
style_rows(ws, 5, 19, 4)

# Colorear outliers
for r in [17, 19]:
    for c in range(1, 5):
        ws.cell(row=r, column=c).fill = red_bg
        ws.cell(row=r, column=c).font = rfont

# Colorear los puntuales
for r in [5, 8, 9, 10, 12, 15, 16, 18]:
    ws.cell(row=r, column=4).font = gfont

# ─── BLOQUE EXPLICATIVO ───
r = 22
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1, value="¿QUÉ NOS DICEN LOS NÚMEROS?").font = sfont

# Promedio
r = 24
ws.cell(row=r, column=1, value="PROMEDIO").font = blfont
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="=AVERAGE(C5:C19)")
ws.cell(row=r, column=2).font = blfont
ws.cell(row=r, column=2).number_format = "0.0"
ws.cell(row=r, column=2).alignment = ctr
ws.cell(row=r, column=2).border = border
ws.merge_cells(f"C{r}:D{r}")
ws.cell(row=r, column=3, value="Sumar todos los valores y dividir por 15. Incluye los outliers.").font = note
ws.cell(row=r, column=3).alignment = wrap
ws.cell(row=r, column=3).border = border

r = 25
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = "→ Da alrededor de 5.5 min. Pero, ¿es real? La mayoría llegó con 0 a 4 min de demora. Los dos viajes de 25 y 35 min arrastran el promedio hacia arriba."
ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="334155")
ws.cell(row=r, column=1).alignment = wrap
ws.row_dimensions[r].height = 35

# Mediana
r = 27
ws.cell(row=r, column=1, value="MEDIANA").font = blfont
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="=MEDIAN(C5:C19)")
ws.cell(row=r, column=2).font = blfont
ws.cell(row=r, column=2).number_format = "0.0"
ws.cell(row=r, column=2).alignment = ctr
ws.cell(row=r, column=2).border = border
ws.merge_cells(f"C{r}:D{r}")
ws.cell(row=r, column=3, value="Ordenar todos los valores y tomar el del medio.").font = note
ws.cell(row=r, column=3).alignment = wrap
ws.cell(row=r, column=3).border = border

r = 28
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = "→ Da 2 min. Esto refleja mucho mejor la realidad: la mayoría de los colectivos llega con ~2 min de demora. Los outliers no la afectan."
ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="334155")
ws.cell(row=r, column=1).alignment = wrap
ws.row_dimensions[r].height = 35

# Moda
r = 30
ws.cell(row=r, column=1, value="MODA").font = blfont
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="=MODE(C5:C19)")
ws.cell(row=r, column=2).font = blfont
ws.cell(row=r, column=2).number_format = "0"
ws.cell(row=r, column=2).alignment = ctr
ws.cell(row=r, column=2).border = border
ws.merge_cells(f"C{r}:D{r}")
ws.cell(row=r, column=3, value="El valor que más se repite.").font = note
ws.cell(row=r, column=3).alignment = wrap
ws.cell(row=r, column=3).border = border

r = 31
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = "→ Da 2 min. Es la demora más frecuente. Aparece 6 veces en 15 llegadas."
ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="334155")
ws.cell(row=r, column=1).alignment = wrap

# Separador
r = 33
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1, value="¿QUÉ TAN DISPERSOS ESTÁN LOS DATOS?").font = sfont

# Rango
r = 35
ws.cell(row=r, column=1, value="RANGO").font = ofont
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="=MAX(C5:C19)-MIN(C5:C19)")
ws.cell(row=r, column=2).font = ofont
ws.cell(row=r, column=2).number_format = "0"
ws.cell(row=r, column=2).alignment = ctr
ws.cell(row=r, column=2).border = border
ws.merge_cells(f"C{r}:D{r}")
ws.cell(row=r, column=3, value="Máximo menos mínimo. ¿Cuánta diferencia hay entre el mejor y el peor?").font = note
ws.cell(row=r, column=3).alignment = wrap
ws.cell(row=r, column=3).border = border

r = 36
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = "→ Da 36 min (de -1 a 35). Un rango enorme. Hay colectivos que llegan antes y otros que llegan 35 min tarde."
ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="334155")
ws.cell(row=r, column=1).alignment = wrap
ws.row_dimensions[r].height = 30

# Desvío
r = 38
ws.cell(row=r, column=1, value="DESVÍO ESTÁNDAR").font = ofont
ws.cell(row=r, column=1).border = border
ws.cell(row=r, column=2, value="=STDEV(C5:C19)")
ws.cell(row=r, column=2).font = ofont
ws.cell(row=r, column=2).number_format = "0.0"
ws.cell(row=r, column=2).alignment = ctr
ws.cell(row=r, column=2).border = border
ws.merge_cells(f"C{r}:D{r}")
ws.cell(row=r, column=3, value="Cuánto se alejan los datos del promedio, en general.").font = note
ws.cell(row=r, column=3).alignment = wrap
ws.cell(row=r, column=3).border = border

r = 39
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = "→ Desvío alto = los datos están muy dispersos. Desvío bajo = los datos son parecidos entre sí."
ws.cell(row=r, column=1).font = Font(name="Calibri", size=10, color="334155")
ws.cell(row=r, column=1).alignment = wrap

# ─── CONCLUSIÓN FINAL ───
r = 42
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = "RESUMEN: ¿QUÉ USO PARA QUÉ?"
ws.cell(row=r, column=1).font = Font(name="Calibri", size=12, bold=True, color=NAVY)

resumen = [
    ("¿Cuál es el valor típico?", "MEDIANA", "Más confiable cuando hay valores extremos"),
    ("¿Cuál es el más frecuente?", "MODA", "El valor que más se repite"),
    ("¿Cuál es el promedio matemático?", "PROMEDIO", "Útil si los datos son parejos, se distorsiona con extremos"),
    ("¿Cuánta diferencia hay?", "RANGO", "Diferencia entre el máximo y el mínimo"),
    ("¿Qué tan dispersos están?", "DESVÍO ESTÁNDAR", "Bajo = estable, Alto = impredecible"),
]

for i, h in enumerate(["Pregunta", "Medida", "Cuándo sirve"], 1):
    ws.cell(row=43, column=i, value=h)
header(ws, 43, 3)

for i, (preg, medida, cuando) in enumerate(resumen, 44):
    ws.cell(row=i, column=1, value=preg).font = bfont
    ws.cell(row=i, column=1).border = border
    ws.cell(row=i, column=1).alignment = wrap
    ws.cell(row=i, column=2, value=medida).font = blfont
    ws.cell(row=i, column=2).border = border
    ws.cell(row=i, column=2).alignment = ctr
    ws.cell(row=i, column=3, value=cuando).font = nfont
    ws.cell(row=i, column=3).border = border
    ws.cell(row=i, column=3).alignment = wrap
    if (i - 44) % 2 == 1:
        for c in range(1, 4):
            ws.cell(row=i, column=c).fill = alt

# Mensaje final
r = 50
ws.merge_cells(f"A{r}:D{r}")
ws.cell(row=r, column=1).value = (
    "EL PROMEDIO SOLO NO ALCANZA. Siempre mirá también la mediana y el desvío. "
    "Si el promedio y la mediana son muy diferentes, hay outliers. "
    "Si el desvío es alto, el proceso es impredecible."
)
ws.cell(row=r, column=1).font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
ws.cell(row=r, column=1).alignment = wrap
ws.row_dimensions[r].height = 45

widths(ws, [22, 16, 18, 44])

# ═══════════════════════════════════════════════════
# HOJA 2: "Puntual vs Impredecible"
# Mismo promedio, distinta variabilidad
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("Puntual vs Impredecible")
ws2.sheet_properties.tabColor = ORANGE

ws2.merge_cells("A1:C1")
ws2["A1"].value = "¿Cumplir el promedio es suficiente?"
ws2["A1"].font = tfont

ws2.merge_cells("A2:C2")
ws2["A2"].value = "Dos choferes, mismo promedio de demora (~3 min), pero uno es puntual y el otro es impredecible."
ws2["A2"].font = note
ws2.row_dimensions[2].height = 28

for i, h in enumerate(["Viaje #", "Pérez (Puntual)", "Gómez (Impredecible)"], 1):
    ws2.cell(row=4, column=i, value=h)
ws2.cell(row=4, column=2).fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
ws2.cell(row=4, column=2).font = hfont
ws2.cell(row=4, column=3).fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
ws2.cell(row=4, column=3).font = hfont
ws2.cell(row=4, column=1).fill = hfill
ws2.cell(row=4, column=1).font = hfont
for c in range(1, 4):
    ws2.cell(row=4, column=c).alignment = ctr
    ws2.cell(row=4, column=c).border = border

# Pérez: siempre entre 2 y 4 min (predecible)
# Gómez: va de -5 a 15 min (caótico, pero promedio ~3)
perez =  [3, 2, 4, 3, 2, 3, 4, 3, 2, 3, 4, 3, 2, 3, 3]
gomez =  [-2, 8, 0, 12, 1, -5, 15, 2, 10, -3, 7, 1, 14, -4, 0]

for i in range(15):
    r = 5 + i
    ws2.cell(row=r, column=1, value=i + 1)
    ws2.cell(row=r, column=2, value=perez[i])
    ws2.cell(row=r, column=3, value=gomez[i])
style_rows(ws2, 5, 19, 3)

# Colorear extremos de Gómez
for i, v in enumerate(gomez):
    r = 5 + i
    if v >= 10 or v <= -4:
        ws2.cell(row=r, column=3).fill = red_bg
        ws2.cell(row=r, column=3).font = rfont

# Estadísticas
r = 22
ws2.cell(row=r, column=1, value="COMPARACIÓN").font = sfont

labels = ["Promedio", "Mediana", "Moda", "Rango", "Desvío estándar", "Mínimo", "Máximo"]
formulas_p = [
    "=AVERAGE(B5:B19)", "=MEDIAN(B5:B19)", "=MODE(B5:B19)",
    "=MAX(B5:B19)-MIN(B5:B19)", "=STDEV(B5:B19)", "=MIN(B5:B19)", "=MAX(B5:B19)"
]
formulas_g = [
    "=AVERAGE(C5:C19)", "=MEDIAN(C5:C19)", "=MODE(C5:C19)",
    "=MAX(C5:C19)-MIN(C5:C19)", "=STDEV(C5:C19)", "=MIN(C5:C19)", "=MAX(C5:C19)"
]

ws2.cell(row=23, column=2, value="Pérez").font = gfont
ws2.cell(row=23, column=2).alignment = ctr
ws2.cell(row=23, column=3, value="Gómez").font = ofont
ws2.cell(row=23, column=3).alignment = ctr

for i, (label, fp, fg) in enumerate(zip(labels, formulas_p, formulas_g)):
    r = 24 + i
    ws2.cell(row=r, column=1, value=label).font = bfont
    ws2.cell(row=r, column=1).border = border
    ws2.cell(row=r, column=2, value=fp)
    ws2.cell(row=r, column=2).font = gfont
    ws2.cell(row=r, column=2).number_format = "0.0"
    ws2.cell(row=r, column=2).alignment = ctr
    ws2.cell(row=r, column=2).border = border
    ws2.cell(row=r, column=3, value=fg)
    ws2.cell(row=r, column=3).font = ofont
    ws2.cell(row=r, column=3).number_format = "0.0"
    ws2.cell(row=r, column=3).alignment = ctr
    ws2.cell(row=r, column=3).border = border

r = 33
ws2.merge_cells(f"A{r}:C{r}")
ws2.cell(row=r, column=1).value = (
    "Pérez: promedio 3 min, desvío 0.7 → Siempre llega entre 2 y 4 min. Es predecible. Podés confiar."
)
ws2.cell(row=r, column=1).font = gfont
ws2.cell(row=r, column=1).alignment = wrap
ws2.cell(row=r, column=1).fill = green_bg
ws2.row_dimensions[r].height = 30

r = 34
ws2.merge_cells(f"A{r}:C{r}")
ws2.cell(row=r, column=1).value = (
    "Gómez: promedio ~3 min, desvío ~7 → Un día llega 5 min antes, otro llega 15 min tarde. Impredecible. No podés planificar."
)
ws2.cell(row=r, column=1).font = ofont
ws2.cell(row=r, column=1).alignment = wrap
ws2.cell(row=r, column=1).fill = orange_bg
ws2.row_dimensions[r].height = 30

r = 36
ws2.merge_cells(f"A{r}:C{r}")
ws2.cell(row=r, column=1).value = (
    "MORALEJA: Dos personas pueden tener el mismo promedio, "
    "pero una es confiable y la otra es un problema. "
    "El promedio solo no alcanza para entender qué pasa. "
    "Siempre hay que mirar la variabilidad."
)
ws2.cell(row=r, column=1).font = Font(name="Calibri", size=12, bold=True, color=NAVY)
ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[r].height = 55

widths(ws2, [14, 22, 22])

# ─── Guardar ───
output = "/home/branko007/projects/iselin_training/tools/estadisticas_iselin_v2.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
