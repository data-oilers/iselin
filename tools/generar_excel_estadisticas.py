#!/usr/bin/env python3
"""
Genera un Excel con datos de ejemplo de Autotransportes ISELIN
para explicar conceptos estadísticos: promedio, mediana, moda,
rango, desviación estándar, y variabilidad vs. promedio.

Compatible con Excel 2016+.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─── Estilos ───
NAVY = "0F172A"
BLUE = "3B82F6"
TEAL = "0D9488"
ORANGE = "EA580C"
GREEN = "059669"
WHITE = "FFFFFF"
LIGHT_GRAY = "F1F5F9"
MID_GRAY = "E2E8F0"

header_font = Font(name="Calibri", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_font = Font(name="Calibri", size=14, bold=True, color=NAVY)
subtitle_font = Font(name="Calibri", size=11, bold=True, color=BLUE)
normal_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", size=11, bold=True)
stat_font = Font(name="Calibri", size=11, bold=True, color=BLUE)
stat_label_font = Font(name="Calibri", size=11, bold=True, color=NAVY)
green_font = Font(name="Calibri", size=11, bold=True, color=GREEN)
orange_font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
note_font = Font(name="Calibri", size=10, italic=True, color="64748B")
thin_border = Border(
    left=Side(style="thin", color=MID_GRAY),
    right=Side(style="thin", color=MID_GRAY),
    top=Side(style="thin", color=MID_GRAY),
    bottom=Side(style="thin", color=MID_GRAY),
)
center = Alignment(horizontal="center", vertical="center")
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border


def style_data_range(ws, start_row, end_row, cols):
    alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center
            if (row - start_row) % 2 == 1:
                cell.fill = alt_fill


def auto_width(ws, cols, min_width=12):
    for col in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = min_width


# ═══════════════════════════════════════
# CREAR WORKBOOK
# ═══════════════════════════════════════
wb = openpyxl.Workbook()

# ─── HOJA 1: Consumo de Gasoil por Colectivo ───
ws1 = wb.active
ws1.title = "Consumo Gasoil"
ws1.sheet_properties.tabColor = BLUE

# Título
ws1.merge_cells("A1:F1")
ws1["A1"].value = "Consumo de Gasoil por Colectivo — Semana del 14 al 18 de Abril 2026"
ws1["A1"].font = title_font
ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")

ws1.merge_cells("A2:F2")
ws1["A2"].value = "Litros consumidos por día. ¿Todos los colectivos consumen parecido o hay mucha diferencia?"
ws1["A2"].font = note_font

# Headers
headers1 = ["Colectivo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 6)

# Datos: 10 colectivos, consumo diario en litros
# Diseñados para que el PROMEDIO sea similar (~48 lts) pero la VARIABILIDAD sea muy distinta
data_gasoil = [
    # Colectivo, Lun, Mar, Mie, Jue, Vie
    ["INT-101", 47, 49, 48, 50, 46],      # Muy estable
    ["INT-102", 46, 48, 50, 47, 49],      # Muy estable
    ["INT-103", 45, 51, 47, 49, 48],      # Estable
    ["INT-104", 30, 65, 35, 70, 40],      # MUY variable
    ["INT-105", 48, 47, 49, 48, 48],      # Muy estable
    ["INT-106", 25, 72, 28, 68, 47],      # MUY variable
    ["INT-107", 50, 46, 48, 49, 47],      # Estable
    ["INT-108", 44, 52, 46, 50, 48],      # Estable
    ["INT-109", 35, 60, 40, 55, 50],      # Variable
    ["INT-110", 48, 48, 48, 48, 48],      # Perfectamente estable
]

for i, row in enumerate(data_gasoil, 5):
    for j, val in enumerate(row):
        ws1.cell(row=i, column=j + 1, value=val)
style_data_range(ws1, 5, 14, 6)

# Fórmulas de resumen por colectivo
ws1.cell(row=4, column=8, value="Promedio").font = header_font
ws1.cell(row=4, column=8).fill = header_fill
ws1.cell(row=4, column=8).alignment = center
ws1.cell(row=4, column=8).border = thin_border

ws1.cell(row=4, column=9, value="Desvío Estándar").font = header_font
ws1.cell(row=4, column=9).fill = header_fill
ws1.cell(row=4, column=9).alignment = center
ws1.cell(row=4, column=9).border = thin_border

ws1.cell(row=4, column=10, value="Rango").font = header_font
ws1.cell(row=4, column=10).fill = header_fill
ws1.cell(row=4, column=10).alignment = center
ws1.cell(row=4, column=10).border = thin_border

# Columna de interpretación
ws1.cell(row=4, column=11, value="Interpretación").font = header_font
ws1.cell(row=4, column=11).fill = header_fill
ws1.cell(row=4, column=11).alignment = center
ws1.cell(row=4, column=11).border = thin_border

interpretaciones = [
    "Estable. Consumo predecible, sin sorpresas.",
    "Estable. Valores muy parejos entre días.",
    "Estable. Variación mínima, proceso bajo control.",
    "ATENCIÓN: Muy variable. Un día 30 lts, otro 70. Revisar vehículo o chofer.",
    "Estable. Casi idéntico todos los días.",
    "ATENCIÓN: Muy variable. Oscila entre 25 y 72 lts. Causa a investigar.",
    "Estable. Consumo parejo toda la semana.",
    "Estable. Leve variación, dentro de lo esperable.",
    "Variable. Oscila bastante, pero menos que INT-104/106.",
    "Perfectamente estable. Mismo consumo todos los días.",
]

for i in range(5, 15):
    rng = f"B{i}:F{i}"
    ws1.cell(row=i, column=8).value = f"=AVERAGE({rng})"
    ws1.cell(row=i, column=8).number_format = "0.0"
    ws1.cell(row=i, column=8).font = bold_font
    ws1.cell(row=i, column=8).alignment = center
    ws1.cell(row=i, column=8).border = thin_border

    ws1.cell(row=i, column=9).value = f"=STDEV({rng})"
    ws1.cell(row=i, column=9).number_format = "0.0"
    ws1.cell(row=i, column=9).alignment = center
    ws1.cell(row=i, column=9).border = thin_border

    ws1.cell(row=i, column=10).value = f"=MAX({rng})-MIN({rng})"
    ws1.cell(row=i, column=10).number_format = "0"
    ws1.cell(row=i, column=10).alignment = center
    ws1.cell(row=i, column=10).border = thin_border

    interp = interpretaciones[i - 5]
    ws1.cell(row=i, column=11).value = interp
    ws1.cell(row=i, column=11).alignment = left_wrap
    ws1.cell(row=i, column=11).border = thin_border
    if "ATENCIÓN" in interp:
        ws1.cell(row=i, column=11).font = Font(name="Calibri", size=10, bold=True, color=ORANGE)
    else:
        ws1.cell(row=i, column=11).font = Font(name="Calibri", size=10, color="059669")

# Nota explicativa de columnas
row_nota = 16
ws1.merge_cells(f"H{row_nota}:K{row_nota}")
ws1.cell(row=row_nota, column=8).value = (
    "Promedio = valor típico | Desvío = cuánto varía | Rango = diferencia entre máximo y mínimo | "
    "Si el desvío es bajo (<3), el colectivo es estable. Si es alto (>10), hay un problema."
)
ws1.cell(row=row_nota, column=8).font = note_font
ws1.cell(row=row_nota, column=8).alignment = left_wrap
ws1.row_dimensions[row_nota].height = 35

# Estadísticas generales
row_stats = 18
ws1.cell(row=row_stats, column=1, value="ESTADÍSTICAS GENERALES").font = subtitle_font
ws1.merge_cells(f"A{row_stats}:C{row_stats}")

stats = [
    ("Promedio general (lts/día)", f"=AVERAGE(B5:F14)",
     "Sumar todos los valores y dividir por la cantidad. Es el valor 'típico', pero se distorsiona si hay valores extremos."),
    ("Mediana", f"=MEDIAN(B5:F14)",
     "Ordenar todos los valores y tomar el del medio. No se ve afectada por extremos. Si es muy diferente al promedio, hay outliers."),
    ("Moda", f"=MODE(B5:F14)",
     "El valor que más veces aparece. Indica cuál es el consumo más frecuente entre todos los registros."),
    ("Rango general", f"=MAX(B5:F14)-MIN(B5:F14)",
     "Máximo menos mínimo. Cuanto más grande, más diferencia hay entre el mejor y el peor caso."),
    ("Desvío estándar general", f"=STDEV(B5:F14)",
     "Mide cuánto se alejan los datos del promedio. Desvío bajo (<5) = datos parejos. Desvío alto (>10) = mucha dispersión."),
    ("Mínimo", f"=MIN(B5:F14)",
     "El consumo más bajo registrado en toda la semana para todos los colectivos."),
    ("Máximo", f"=MAX(B5:F14)",
     "El consumo más alto registrado. Si es mucho mayor que el promedio, hay que investigar por qué."),
]

for i, (label, formula, desc) in enumerate(stats):
    r = row_stats + 1 + i
    ws1.cell(row=r, column=1, value=label).font = stat_label_font
    ws1.cell(row=r, column=1).border = thin_border
    ws1.cell(row=r, column=2, value=formula)
    ws1.cell(row=r, column=2).font = stat_font
    ws1.cell(row=r, column=2).number_format = "0.0"
    ws1.cell(row=r, column=2).alignment = center
    ws1.cell(row=r, column=2).border = thin_border
    ws1.cell(row=r, column=3, value=desc).font = note_font
    ws1.cell(row=r, column=3).alignment = left_wrap
    ws1.cell(row=r, column=3).border = thin_border

# Conclusión
row_conc = row_stats + len(stats) + 2
ws1.merge_cells(f"A{row_conc}:K{row_conc}")
ws1.cell(row=row_conc, column=1).value = (
    "CONCLUSIÓN CLAVE: El INT-104 y el INT-106 tienen un promedio similar al resto (~48 lts), "
    "pero su variabilidad es enorme (desvío > 15). Esto indica un problema: "
    "el promedio parece correcto, pero el proceso es inestable. "
    "Un colectivo que un día consume 25 lts y otro día 72 lts necesita revisión, "
    "aunque su promedio sea 'normal'."
)
ws1.cell(row=row_conc, column=1).font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
ws1.cell(row=row_conc, column=1).alignment = left_wrap
ws1.row_dimensions[row_conc].height = 50

# Nota pedagógica adicional
row_ped = row_conc + 2
ws1.merge_cells(f"A{row_ped}:K{row_ped}")
ws1.cell(row=row_ped, column=1).value = (
    "¿CÓMO LEER ESTA TABLA? → Mirá primero la columna 'Promedio': te dice cuánto consume en promedio cada colectivo. "
    "Después mirá 'Desvío Estándar': si es bajo (1-3), el consumo es parejo y predecible. Si es alto (>10), hay días muy altos y muy bajos. "
    "Por último, mirá 'Rango': te dice la diferencia entre el peor día y el mejor día. Un rango de 4 es normal. Un rango de 47 es una alarma."
)
ws1.cell(row=row_ped, column=1).font = Font(name="Calibri", size=10, italic=True, color="334155")
ws1.cell(row=row_ped, column=1).alignment = left_wrap
ws1.row_dimensions[row_ped].height = 55

row_msg = row_ped + 2
ws1.merge_cells(f"A{row_msg}:K{row_msg}")
ws1.cell(row=row_msg, column=1).value = "EL PROMEDIO SOLO NO ALCANZA. Siempre hay que mirar la variabilidad."
ws1.cell(row=row_msg, column=1).font = Font(name="Calibri", size=12, bold=True, color=NAVY)
ws1.cell(row=row_msg, column=1).alignment = Alignment(horizontal="center", vertical="center")

auto_width(ws1, 11, 15)
ws1.column_dimensions["A"].width = 26
ws1.column_dimensions["C"].width = 48
ws1.column_dimensions["H"].width = 14
ws1.column_dimensions["I"].width = 18
ws1.column_dimensions["J"].width = 12
ws1.column_dimensions["K"].width = 52


# ─── HOJA 2: Tiempos de Viaje por Ruta ───
ws2 = wb.create_sheet("Tiempos de Viaje")
ws2.sheet_properties.tabColor = TEAL

ws2.merge_cells("A1:E1")
ws2["A1"].value = "Tiempos de Viaje Ruta 5 — Últimos 20 viajes (minutos)"
ws2["A1"].font = title_font

ws2.merge_cells("A2:E2")
ws2["A2"].value = "¿Cuánto tarda normalmente un viaje? ¿Hay viajes que se desvían mucho del tiempo esperado?"
ws2["A2"].font = note_font

headers2 = ["Viaje #", "Chofer", "Tiempo (min)", "Día", "Horario"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 5)

# Datos diseñados para mostrar mediana vs promedio con outliers
tiempos = [
    (1, "García", 42, "Lunes", "Mañana"),
    (2, "López", 45, "Lunes", "Tarde"),
    (3, "Martínez", 43, "Martes", "Mañana"),
    (4, "García", 44, "Martes", "Tarde"),
    (5, "Rodríguez", 41, "Miércoles", "Mañana"),
    (6, "López", 43, "Miércoles", "Tarde"),
    (7, "Fernández", 46, "Jueves", "Mañana"),
    (8, "García", 42, "Jueves", "Tarde"),
    (9, "Martínez", 44, "Viernes", "Mañana"),
    (10, "Rodríguez", 43, "Viernes", "Tarde"),
    (11, "López", 45, "Lunes", "Mañana"),
    (12, "Fernández", 42, "Lunes", "Tarde"),
    (13, "García", 43, "Martes", "Mañana"),
    (14, "Martínez", 44, "Martes", "Tarde"),
    (15, "Rodríguez", 85, "Miércoles", "Mañana"),  # OUTLIER: ruta cortada
    (16, "López", 42, "Miércoles", "Tarde"),
    (17, "García", 43, "Jueves", "Mañana"),
    (18, "Fernández", 44, "Jueves", "Tarde"),
    (19, "Martínez", 90, "Viernes", "Mañana"),  # OUTLIER: accidente en ruta
    (20, "Rodríguez", 43, "Viernes", "Tarde"),
]

for i, (num, chofer, tiempo, dia, horario) in enumerate(tiempos, 5):
    ws2.cell(row=i, column=1, value=num)
    ws2.cell(row=i, column=2, value=chofer)
    ws2.cell(row=i, column=3, value=tiempo)
    ws2.cell(row=i, column=4, value=dia)
    ws2.cell(row=i, column=5, value=horario)
style_data_range(ws2, 5, 24, 5)

# Resaltar outliers
for r in [19, 23]:  # Filas de los viajes 15 y 19
    for c in range(1, 6):
        ws2.cell(row=r, column=c).fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
        ws2.cell(row=r, column=c).font = Font(name="Calibri", size=11, bold=True, color="DC2626")

# Estadísticas
row_s2 = 27
ws2.cell(row=row_s2, column=1, value="ESTADÍSTICAS").font = subtitle_font

# Nota sobre outliers
row_out = 26
ws2.merge_cells(f"A{row_out}:E{row_out}")
ws2.cell(row=row_out, column=1).value = (
    "Las filas en rojo son OUTLIERS (valores atípicos): el viaje 15 tardó 85 min por ruta cortada y el viaje 19 tardó 90 min por un accidente. "
    "Estos 2 valores no representan lo normal, pero afectan al promedio. Observá la diferencia entre promedio y mediana."
)
ws2.cell(row=row_out, column=1).font = Font(name="Calibri", size=10, italic=True, color="DC2626")
ws2.cell(row=row_out, column=1).alignment = left_wrap
ws2.row_dimensions[row_out].height = 40

row_s2 = 28

stats2 = [
    ("Promedio", "=AVERAGE(C5:C24)",
     "Suma todos los tiempos y divide por 20. INCLUYE los outliers, por eso da más alto de lo esperado."),
    ("Mediana", "=MEDIAN(C5:C24)",
     "Ordena los 20 tiempos y toma el del medio. NO se ve afectada por los 2 viajes extremos. Mucho más representativa."),
    ("Moda", "=MODE(C5:C24)",
     "El tiempo que más veces se repite. Indica cuál es la duración más habitual de un viaje."),
    ("Rango", "=MAX(C5:C24)-MIN(C5:C24)",
     "Diferencia entre el viaje más largo (90 min) y el más corto (41 min). Un rango de ~49 min es enorme."),
    ("Desvío estándar", "=STDEV(C5:C24)",
     "Cuánto se dispersan los tiempos. Un desvío alto indica que hay viajes con tiempos muy diferentes entre sí."),
    ("Mínimo", "=MIN(C5:C24)",
     "El viaje más rápido. Representa el mejor escenario posible en esta ruta."),
    ("Máximo", "=MAX(C5:C24)",
     "El viaje más lento. Acá debemos preguntarnos: ¿por qué tardó tanto? ¿Es algo que se puede prevenir?"),
]

for i, (label, formula, desc) in enumerate(stats2):
    r = row_s2 + 1 + i
    ws2.cell(row=r, column=1, value=label).font = stat_label_font
    ws2.cell(row=r, column=1).border = thin_border
    ws2.cell(row=r, column=2, value=formula)
    ws2.cell(row=r, column=2).font = stat_font
    ws2.cell(row=r, column=2).number_format = "0.0"
    ws2.cell(row=r, column=2).alignment = center
    ws2.cell(row=r, column=2).border = thin_border
    ws2.cell(row=r, column=3, value=desc).font = note_font
    ws2.cell(row=r, column=3).alignment = left_wrap
    ws2.cell(row=r, column=3).border = thin_border

row_conc2 = row_s2 + len(stats2) + 2
ws2.merge_cells(f"A{row_conc2}:E{row_conc2}")
ws2.cell(row=row_conc2, column=1).value = (
    "CONCLUSIÓN: El promedio da ~47 min, pero la mediana da ~43 min. "
    "¿Por qué la diferencia? Porque 2 viajes (el 15 y el 19) tardaron 85 y 90 minutos. "
    "Esos valores extremos arrastran el promedio hacia arriba. "
    "La mediana es más confiable en estos casos: el viaje típico de la Ruta 5 tarda unos 43 minutos."
)
ws2.cell(row=row_conc2, column=1).font = Font(name="Calibri", size=11, bold=True, color=TEAL)
ws2.cell(row=row_conc2, column=1).alignment = left_wrap
ws2.row_dimensions[row_conc2].height = 50

# Nota pedagógica
row_ped2 = row_conc2 + 2
ws2.merge_cells(f"A{row_ped2}:E{row_ped2}")
ws2.cell(row=row_ped2, column=1).value = (
    "PARA RECORDAR: Cuando el promedio y la mediana dan valores muy distintos, es señal de que hay outliers (valores extremos). "
    "En esos casos, la mediana refleja mejor la realidad. El promedio es útil cuando los datos son parejos y no hay valores raros."
)
ws2.cell(row=row_ped2, column=1).font = Font(name="Calibri", size=10, italic=True, color="334155")
ws2.cell(row=row_ped2, column=1).alignment = left_wrap
ws2.row_dimensions[row_ped2].height = 45

auto_width(ws2, 5, 14)
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["C"].width = 48
ws2.column_dimensions["E"].width = 48
ws2.column_dimensions["E"].width = 42


# ─── HOJA 3: El problema de la variabilidad ───
ws3 = wb.create_sheet("Variabilidad")
ws3.sheet_properties.tabColor = ORANGE

ws3.merge_cells("A1:D1")
ws3["A1"].value = "¿Cumplir el promedio es suficiente?"
ws3["A1"].font = title_font

ws3.merge_cells("A2:D2")
ws3["A2"].value = "Dos colectivos con el MISMO promedio de consumo, pero resultados muy diferentes"
ws3["A2"].font = note_font

# Colectivo A: estable
ws3.cell(row=4, column=1, value="Día").font = header_font
ws3.cell(row=4, column=1).fill = header_fill
ws3.cell(row=4, column=1).alignment = center
ws3.cell(row=4, column=1).border = thin_border

ws3.cell(row=4, column=2, value="INT-101 (Estable)").font = header_font
ws3.cell(row=4, column=2).fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
ws3.cell(row=4, column=2).alignment = center
ws3.cell(row=4, column=2).border = thin_border

ws3.cell(row=4, column=3, value="INT-104 (Variable)").font = header_font
ws3.cell(row=4, column=3).fill = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
ws3.cell(row=4, column=3).alignment = center
ws3.cell(row=4, column=3).border = thin_border

ws3.cell(row=4, column=4, value="Meta (lts/día)").font = header_font
ws3.cell(row=4, column=4).fill = header_fill
ws3.cell(row=4, column=4).alignment = center
ws3.cell(row=4, column=4).border = thin_border

dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes",
        "Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
estable = [48, 49, 47, 50, 46, 48, 51, 47, 49, 48]
variable = [25, 72, 30, 68, 45, 28, 70, 35, 65, 42]
meta = [48] * 10

for i, (d, e, v, m) in enumerate(zip(dias, estable, variable, meta), 5):
    ws3.cell(row=i, column=1, value=d)
    ws3.cell(row=i, column=2, value=e)
    ws3.cell(row=i, column=3, value=v)
    ws3.cell(row=i, column=4, value=m)
style_data_range(ws3, 5, 14, 4)

# Estadísticas comparativas
row_s3 = 17
ws3.cell(row=row_s3, column=1, value="COMPARACIÓN").font = subtitle_font

comp_stats = [
    ("Promedio", "=AVERAGE(B5:B14)", "=AVERAGE(C5:C14)"),
    ("Mediana", "=MEDIAN(B5:B14)", "=MEDIAN(C5:C14)"),
    ("Desvío estándar", "=STDEV(B5:B14)", "=STDEV(C5:C14)"),
    ("Rango", "=MAX(B5:B14)-MIN(B5:B14)", "=MAX(C5:C14)-MIN(C5:C14)"),
    ("Mínimo", "=MIN(B5:B14)", "=MIN(C5:C14)"),
    ("Máximo", "=MAX(B5:B14)", "=MAX(C5:C14)"),
]

ws3.cell(row=row_s3 + 1, column=2, value="INT-101").font = green_font
ws3.cell(row=row_s3 + 1, column=2).alignment = center
ws3.cell(row=row_s3 + 1, column=3, value="INT-104").font = orange_font
ws3.cell(row=row_s3 + 1, column=3).alignment = center

for i, (label, f1, f2) in enumerate(comp_stats):
    r = row_s3 + 2 + i
    ws3.cell(row=r, column=1, value=label).font = stat_label_font
    ws3.cell(row=r, column=1).border = thin_border
    ws3.cell(row=r, column=2, value=f1)
    ws3.cell(row=r, column=2).font = green_font
    ws3.cell(row=r, column=2).number_format = "0.0"
    ws3.cell(row=r, column=2).alignment = center
    ws3.cell(row=r, column=2).border = thin_border
    ws3.cell(row=r, column=3, value=f2)
    ws3.cell(row=r, column=3).font = orange_font
    ws3.cell(row=r, column=3).number_format = "0.0"
    ws3.cell(row=r, column=3).alignment = center
    ws3.cell(row=r, column=3).border = thin_border

row_conc3 = row_s3 + len(comp_stats) + 4
ws3.merge_cells(f"A{row_conc3}:D{row_conc3}")
ws3.cell(row=row_conc3, column=1).value = (
    "CONCLUSIÓN: Ambos colectivos tienen casi el mismo promedio (~48 lts). "
    "Pero el INT-101 es predecible y estable (desvío ~1.4), mientras que el INT-104 es un caos (desvío ~18). "
    "Cumplir la meta EN PROMEDIO no significa que el proceso esté bajo control. "
    "La variabilidad alta indica que algo anda mal: puede ser el chofer, el estado del vehículo o la ruta."
)
ws3.cell(row=row_conc3, column=1).font = Font(name="Calibri", size=11, bold=True, color=ORANGE)
ws3.cell(row=row_conc3, column=1).alignment = left_wrap
ws3.row_dimensions[row_conc3].height = 55

row_msg = row_conc3 + 2
ws3.merge_cells(f"A{row_msg}:D{row_msg}")
ws3.cell(row=row_msg, column=1).value = (
    "EL PROMEDIO MIENTE SI NO MIRAMOS LA VARIABILIDAD. "
    "Siempre hay que analizar ambas medidas juntas."
)
ws3.cell(row=row_msg, column=1).font = Font(name="Calibri", size=12, bold=True, color=NAVY)
ws3.cell(row=row_msg, column=1).alignment = Alignment(horizontal="center", vertical="center")

auto_width(ws3, 4, 18)
ws3.column_dimensions["A"].width = 22

# ─── Guardar ───
output = "/home/branko007/projects/iselin_training/tools/estadisticas_iselin.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
