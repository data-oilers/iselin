#!/usr/bin/env python3
"""
Genera un Excel con datos simulados de Autotransportes ISELIN
estructurados para cargar en Power BI.

Tablas:
- Viajes (hechos): ~2400 registros, 4 meses (ene-abr 2026)
- Choferes (dimensión): 15 choferes
- Rutas (dimensión): 8 rutas
- Colectivos (dimensión): 20 colectivos
- Calendario (dimensión): ene 2025 a dic 2026

Compatible con Excel 2016+. Cada tabla en su propia hoja.
"""

import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

random.seed(42)

# ─── Estilos ───
NAVY = "0F172A"
WHITE = "FFFFFF"
LIGHT = "F1F5F9"
MID = "E2E8F0"

hfont = Font(name="Calibri", size=11, bold=True, color=WHITE)
hfill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
nfont = Font(name="Calibri", size=11)
border = Border(
    left=Side(style="thin", color=MID), right=Side(style="thin", color=MID),
    top=Side(style="thin", color=MID), bottom=Side(style="thin", color=MID),
)
ctr = Alignment(horizontal="center", vertical="center")
alt_fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")


def write_table(ws, headers, data, tab_color):
    ws.sheet_properties.tabColor = tab_color

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = ctr
        cell.border = border

    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = nfont
            cell.border = border
            cell.alignment = ctr
            if (r - 2) % 2 == 1:
                cell.fill = alt_fill

    for i in range(1, len(headers) + 1):
        max_len = max(len(str(headers[i-1])), 12)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 4

    # Formato de tabla para PBI
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"


# ═══════════════════════════════════════
# DATOS BASE
# ═══════════════════════════════════════

# Choferes
choferes = [
    ("CH001", "García, Juan", "A", 12),
    ("CH002", "López, María", "A", 8),
    ("CH003", "Martínez, Carlos", "B", 5),
    ("CH004", "Rodríguez, Ana", "A", 15),
    ("CH005", "Fernández, Pedro", "B", 3),
    ("CH006", "González, Laura", "A", 10),
    ("CH007", "Díaz, Roberto", "C", 2),
    ("CH008", "Sánchez, Marta", "B", 7),
    ("CH009", "Romero, Diego", "A", 11),
    ("CH010", "Torres, Lucía", "B", 4),
    ("CH011", "Álvarez, Hugo", "C", 1),
    ("CH012", "Ruiz, Silvia", "A", 9),
    ("CH013", "Moreno, Fabián", "B", 6),
    ("CH014", "Castro, Elena", "A", 14),
    ("CH015", "Giménez, Oscar", "C", 2),
]

# Rutas
rutas = [
    ("R001", "Ruta 1 - Centro", "Urbana", 15),
    ("R002", "Ruta 2 - Norte", "Urbana", 18),
    ("R003", "Ruta 3 - Sur", "Urbana", 22),
    ("R004", "Ruta 4 - Este", "Urbana", 20),
    ("R005", "Ruta 5 - San Rafael-Alvear", "Media Distancia", 85),
    ("R006", "Ruta 6 - San Rafael-Malargüe", "Media Distancia", 120),
    ("R007", "Ruta 7 - San Rafael-Mendoza", "Larga Distancia", 230),
    ("R008", "Ruta 8 - San Rafael-Rincón", "Urbana", 12),
]

# Colectivos
colectivos = []
modelos = [
    ("Mercedes-Benz OF 1721", 2018),
    ("Mercedes-Benz OF 1721", 2019),
    ("Mercedes-Benz OF 1724", 2020),
    ("Agrale MA 15.0", 2017),
    ("Agrale MA 15.0", 2019),
    ("Volkswagen 17.230 OD", 2021),
    ("Volkswagen 17.230 OD", 2022),
    ("Mercedes-Benz O-500 M", 2020),
    ("Mercedes-Benz O-500 M", 2021),
    ("Scania K310", 2019),
]

estados = ["Activo", "Activo", "Activo", "Activo", "Activo",
           "Activo", "Activo", "Activo", "En Taller", "Activo",
           "Activo", "Activo", "Activo", "En Taller", "Activo",
           "Activo", "Activo", "Activo", "Activo", "Baja Temporal"]

for i in range(20):
    modelo, anio = modelos[i % len(modelos)]
    colectivos.append((
        f"COL{i+1:03d}",
        f"INT-{100+i+1}",
        modelo,
        anio + (i // 10),
        estados[i],
    ))

# Calendario: ene 2025 a dic 2026
dias_semana_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
meses_es = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

calendario = []
fecha = datetime(2025, 1, 1)
fin = datetime(2026, 12, 31)
while fecha <= fin:
    calendario.append((
        fecha.strftime("%Y-%m-%d"),
        fecha.day,
        meses_es[fecha.month - 1],
        fecha.month,
        fecha.year,
        f"T{(fecha.month - 1) // 3 + 1}",
        fecha.isocalendar()[1],
        dias_semana_es[fecha.weekday()],
        "Sí" if fecha.weekday() < 5 else "No",
    ))
    fecha += timedelta(days=1)

# ═══════════════════════════════════════
# VIAJES (tabla de hechos)
# ═══════════════════════════════════════
viajes = []
viaje_id = 1

# Generar viajes de enero a abril 2026
fecha_inicio = datetime(2026, 1, 1)
fecha_fin = datetime(2026, 4, 30)
fecha = fecha_inicio

while fecha <= fecha_fin:
    # No generar domingos (menos viajes)
    if fecha.weekday() == 6:
        fecha += timedelta(days=1)
        continue

    # Cantidad de viajes por día (más en días hábiles)
    if fecha.weekday() < 5:
        n_viajes = random.randint(18, 28)
    else:
        n_viajes = random.randint(8, 14)

    for _ in range(n_viajes):
        chofer = random.choice(choferes)
        ruta = random.choice(rutas)

        # Solo colectivos activos
        col = random.choice([c for c in colectivos if c[4] == "Activo"])

        # Km base de la ruta con variación
        km_base = ruta[3]
        km = round(km_base * random.uniform(0.92, 1.08), 1)

        # Gasoil proporcional a km con variación por colectivo/chofer
        gasoil_base = km * 0.28  # ~0.28 lts por km
        # Algunos choferes consumen más
        if chofer[0] in ("CH007", "CH011", "CH015"):  # Categoría C
            gasoil_base *= random.uniform(1.10, 1.25)
        gasoil = round(gasoil_base * random.uniform(0.90, 1.15), 1)

        # Pasajeros según tipo de ruta
        if ruta[2] == "Urbana":
            pasajeros = random.randint(25, 65)
        elif ruta[2] == "Media Distancia":
            pasajeros = random.randint(15, 45)
        else:
            pasajeros = random.randint(20, 52)

        # Demora en minutos (-3 a 15, con algunos outliers)
        if random.random() < 0.08:
            demora = random.randint(15, 40)  # Outlier
        elif random.random() < 0.15:
            demora = random.randint(-3, 0)  # Llegó antes
        else:
            demora = random.randint(0, 8)

        # Hora de salida
        hora = random.choice([5,6,6,7,7,7,8,8,9,10,11,12,13,14,15,16,17,18,19,20])
        minuto = random.randint(0, 59)

        # Siniestro (baja probabilidad)
        siniestro = "Sí" if random.random() < 0.02 else "No"

        # Turno
        turno = "Mañana" if hora < 13 else "Tarde"

        viajes.append((
            f"V{viaje_id:05d}",
            fecha.strftime("%Y-%m-%d"),
            chofer[0],
            ruta[0],
            col[0],
            f"{hora:02d}:{minuto:02d}",
            turno,
            km,
            gasoil,
            pasajeros,
            demora,
            siniestro,
        ))
        viaje_id += 1

    fecha += timedelta(days=1)

print(f"Viajes generados: {len(viajes)}")

# ═══════════════════════════════════════
# CREAR EXCEL
# ═══════════════════════════════════════
wb = openpyxl.Workbook()

# Hoja Viajes
ws_v = wb.active
ws_v.title = "Viajes"
write_table(ws_v,
    ["ID_Viaje", "Fecha", "ID_Chofer", "ID_Ruta", "ID_Colectivo",
     "Hora_Salida", "Turno", "Km", "Gasoil_Lts", "Pasajeros", "Demora_Min", "Siniestro"],
    viajes, "3B82F6")

# Hoja Choferes
ws_c = wb.create_sheet("Choferes")
write_table(ws_c,
    ["ID_Chofer", "Nombre", "Categoría", "Antigüedad_Años"],
    choferes, "0D9488")

# Hoja Rutas
ws_r = wb.create_sheet("Rutas")
write_table(ws_r,
    ["ID_Ruta", "Nombre_Ruta", "Tipo", "Km_Base"],
    rutas, "EA580C")

# Hoja Colectivos
ws_col = wb.create_sheet("Colectivos")
write_table(ws_col,
    ["ID_Colectivo", "Interno", "Modelo", "Año", "Estado"],
    colectivos, "059669")

# Hoja Calendario
ws_cal = wb.create_sheet("Calendario")
write_table(ws_cal,
    ["Fecha", "Día", "Mes_Nombre", "Mes_Num", "Año", "Trimestre", "Semana", "Día_Semana", "Es_Hábil"],
    calendario, "7C3AED")

# Guardar
output = "/home/branko007/projects/iselin_training/demo_pbi/datos_iselin.xlsx"
wb.save(output)
print(f"Excel generado: {output}")
print(f"  Viajes: {len(viajes)} registros")
print(f"  Choferes: {len(choferes)}")
print(f"  Rutas: {len(rutas)}")
print(f"  Colectivos: {len(colectivos)}")
print(f"  Calendario: {len(calendario)} días")
